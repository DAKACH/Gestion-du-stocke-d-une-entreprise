import customtkinter as ctk
from customtkinter import CTkImage
from tkinter import ttk
from PIL import Image, ImageTk
from db import Database
from customtkinter import filedialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from faker import Faker
from fpdf import FPDF
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

db = Database("Produis.db")


categories = ["Électronique", "Vêtements", "Meubles"]

mode = "dark"
ctk.set_appearance_mode(mode)


def mode_fonction():
    if sw.get():
        ctk.set_appearance_mode("light")
    else:
        ctk.set_appearance_mode("dark")


def show_heid_fonction():
    if sow_btn._text == "<<":
        sow_btn.configure(text=">>", font=("Arial", 16, "bold"))
        app.geometry("370x700+100+50")
    else:
        sow_btn.configure(text="<<", font=("Arial", 16, "bold"))
        app.geometry("1310x700+100+50")


def add_category():
    new_category = entry_Categorie.get()
    if new_category and new_category not in categories:
        categories.append(new_category)
        ComboBox_Categorie.configure(values=categories)
        entry_Categorie.delete(0, ctk.END)


def resize_image(image_path, size, dark_image_path=None):
    light_image = Image.open(image_path)
    dark_image = Image.open(dark_image_path) if dark_image_path else light_image
    return ctk.CTkImage(light_image=light_image, dark_image=dark_image, size=size)


ctk.set_default_color_theme("green")  # dark-blue green blue


app = ctk.CTk()
app.title("Gestion du Stock App")
app.geometry("1310x700+100+50")
app.resizable(False, False)
app.iconbitmap("images/inventory_management_yWx_icon.ico")

logo_ensate = ctk.CTkImage(
    Image.open("images/project_logo-removebg-preview.png"),
    size=(300, 130),
)
lbl_logo = ctk.CTkLabel(app, text="", image=logo_ensate)
lbl_logo.place(x=1000, y=20)
# ________________________________________________________________________________________
nom = ctk.StringVar()
quantite = ctk.StringVar()
prix = ctk.StringVar()
date = ctk.StringVar()
categorie = ctk.StringVar()
# ____________________________________________________________________________
frame_1 = ctk.CTkFrame(
    app,
    height=680,
    width=350,
    corner_radius=20,
)
frame_1.place(x=10, y=10)
title_1 = ctk.CTkLabel(
    frame_1, text="Gestion du stock d'une entreprise", font=("Arial", 18, "bold")
)
# title_1.place(x=30, y=1)
info_text = (
    "projet 6: gestion du stocke d'une entreprise\n"
    "Superviseur : Pr Aytouna Fouad\n"
    "Développeur : Dakach Otman\n"
    "Email : bitoppt@gmail.com\n"
    "Spécialité : Ingénierie en Cybersécurité et Systèmes Embarqués\n"
    "Date du projet : 7 Janvier 2025\n"
    "Ecole : ENSA TETOUAN"
)

frame_2 = ctk.CTkFrame(
    app,
    height=130,
    width=630,
    corner_radius=20,
)
frame_2.place(x=370, y=10)

ensa_tetouan_image = resize_image("images/ENSATETOUAN.png", (140, 70))
info_text = ctk.CTkLabel(
    frame_2,
    text=info_text,
    font=("Helvetica", 12, "bold"),
    padx=10,
    pady=10,
    anchor="w",
    width=400,
    ## image=ensa_tetouan_image,
    # compound="left",
)
info_text.place(x=100, y=10)

lbl_Name = ctk.CTkLabel(frame_1, text="Nom", font=("Arial", 16))
lbl_Name.place(x=30, y=50)
ent_Name = ctk.CTkEntry(frame_1, textvariable=nom, width=160, font=("Arial", 16))
ent_Name.place(x=150, y=50)

lbl_id = ctk.CTkLabel(frame_1, text="Id", font=("Arial", 16))
lbl_id.place(x=30, y=90)
ent_id = ctk.CTkEntry(frame_1, width=160, font=("Arial", 16))
ent_id.place(x=150, y=90)

lbl_Quantite = ctk.CTkLabel(frame_1, text="Quantité", font=("Arial", 16))
lbl_Quantite.place(x=30, y=130)
ent_Quantite = ctk.CTkEntry(
    frame_1, textvariable=quantite, width=160, font=("Arial", 16)
)
ent_Quantite.place(x=150, y=130)

lbl_Prix = ctk.CTkLabel(frame_1, text="Prix", font=("Arial", 16))
lbl_Prix.place(x=30, y=170)
ent_Prix = ctk.CTkEntry(frame_1, textvariable=prix, width=160, font=("Arial", 16))
ent_Prix.place(x=150, y=170)

lbl_Date = ctk.CTkLabel(frame_1, text="Date", font=("Arial", 16))
lbl_Date.place(x=30, y=210)
ent_Date = ctk.CTkEntry(frame_1, textvariable=date, width=160, font=("Arial", 16))
ent_Date.place(x=150, y=210)

lbl_Categorie = ctk.CTkLabel(frame_1, text="Catégorie", font=("Arial", 16))
lbl_Categorie.place(x=30, y=250)
ComboBox_Categorie = ctk.CTkComboBox(
    frame_1, width=160, font=("Arial", 16), values=categories, state="readonly"
)
ComboBox_Categorie.place(x=150, y=250)

entry_Categorie = ctk.CTkEntry(
    frame_1, textvariable=categorie, placeholder_text="Ajout de catégorie", width=160
)
entry_Categorie.place(x=150, y=290)


add_button_Categorie = ctk.CTkButton(
    frame_1, text="Ajout de catégorie", width=80, height=28, command=add_category
)
add_button_Categorie.place(x=30, y=290)

lbl_Comentair = ctk.CTkLabel(frame_1, text="Comentair", font=("Arial", 16))
lbl_Comentair.place(x=30, y=330)
ent_Comentair = ctk.CTkTextbox(frame_1, width=150, height=100, font=("Arial", 16))
ent_Comentair.place(x=150, y=330)

# ____________________hide show___________________________________
sw = ctk.CTkSwitch(frame_1, text="mode", command=mode_fonction)
sw.place(x=10, y=10)
sow_btn = ctk.CTkButton(
    frame_1,
    text="<<",
    font=("Arial", 16, "bold"),
    width=40,
    cursor="hand2",
    command=show_heid_fonction,
)
sow_btn.place(x=300, y=10)


# __________________________db fonctions______________________________________
def getData(event):
    selected_row = tv.focus()
    data = tv.item(selected_row)
    global row
    row = data["values"]
    if row:
        ent_Name.delete(0, ctk.END)
        ent_Name.insert(0, row[1])  # Nom
        ent_id.delete(0, ctk.END)
        ent_id.insert(0, row[0])  # ID
        ent_Quantite.delete(0, ctk.END)
        ent_Quantite.insert(0, row[2])  # Quantite
        ent_Prix.delete(0, ctk.END)
        ent_Prix.insert(0, row[3])  # Prix
        ent_Date.delete(0, ctk.END)
        ent_Date.insert(0, row[4])  # Date
        ComboBox_Categorie.set(row[5])  # Categorie
        ent_Comentair.delete("1.0", ctk.END)
        ent_Comentair.insert("1.0", row[6] if row[6] else "")  # Commentaire


def displayAll():
    tv.delete(*tv.get_children())
    for row in db.fetch():
        tv.insert("", ctk.END, values=row)


def clear_fields():
    ent_Name.delete(0, ctk.END)
    ent_id.delete(0, ctk.END)
    ent_Quantite.delete(0, ctk.END)
    ent_Prix.delete(0, ctk.END)
    ent_Date.delete(0, ctk.END)
    ComboBox_Categorie.set("")
    ent_Comentair.delete("1.0", ctk.END)


def add_produit():
    if (
        not ent_Name.get()
        or not ent_Quantite.get()
        or not ent_Prix.get()
        or not ent_Date.get()
        or not ComboBox_Categorie.get()
    ):
        show_warning("Tous les champs sont obligatoires")
        return

    try:
        db.insert(
            ent_Name.get(),
            ent_Quantite.get(),
            ent_Prix.get(),
            ent_Date.get(),
            ComboBox_Categorie.get(),
            ent_Comentair.get("1.0", ctk.END).strip(),
        )
        displayAll()
        clear_fields()
        show_success("Produit ajouté avec succès")
    except Exception as e:
        show_error(f"Une erreur: {str(e)}")


def update_produit():
    if not ent_id.get():
        show_warning("Veuillez sélectionner un produit à modifier")
        return

    try:
        db.update(
            ent_id.get(),
            ent_Name.get(),
            ent_Quantite.get(),
            ent_Prix.get(),
            ent_Date.get(),
            ComboBox_Categorie.get(),
            ent_Comentair.get("1.0", ctk.END).strip(),
        )
        displayAll()
        clear_fields()
        show_success("Produit modifié avec succès")
    except Exception as e:
        show_error(f"Une erreur: {str(e)}")


def delete_produit():
    if not ent_id.get():
        show_warning("Veuillez sélectionner un produit à supprimer")
        return

    if show_confirm("Êtes-vous sûr de vouloir supprimer ce produit"):
        try:
            success, message = db.remove(ent_id.get())
            if success:
                displayAll()
                clear_fields()
                show_success(message)
            else:
                show_error(message)
        except Exception as e:
            show_error(f"Une erreur: {str(e)}")


# help functions---------------------------------------------------
def show_warning(message):
    popup = ctk.CTkToplevel(app)
    popup.title("Alerte")
    popup.geometry("530x150+600+300")
    popup.resizable(False, False)
    popup.transient(app)
    popup.grab_set()

    label = ctk.CTkLabel(popup, text=message, font=("Arial", 14))
    label.pack(pady=20)

    ok_button = ctk.CTkButton(popup, text="OK", command=popup.destroy)
    ok_button.pack(pady=10)


def show_success(message):
    popup = ctk.CTkToplevel(app)
    popup.title("Succès")
    popup.geometry("530x150+600+300")
    popup.resizable(False, False)
    popup.transient(app)
    popup.grab_set()

    label = ctk.CTkLabel(popup, text=message, font=("Arial", 14))
    label.pack(pady=20)

    ok_button = ctk.CTkButton(popup, text="OK", command=popup.destroy)
    ok_button.pack(pady=10)


def show_error(message):
    popup = ctk.CTkToplevel(app)
    popup.title("Erreur")
    popup.geometry("530x150+600+300")
    popup.resizable(False, False)
    popup.transient(app)
    popup.grab_set()

    label = ctk.CTkLabel(popup, text=message, font=("Arial", 14))
    label.pack(pady=20)

    ok_button = ctk.CTkButton(popup, text="OK", command=popup.destroy)
    ok_button.pack(pady=10)


def show_confirm(message):
    result = [False]

    popup = ctk.CTkToplevel(app)
    popup.title("Confirmation")
    popup.geometry("530x150+600+300")
    popup.resizable(False, False)
    popup.transient(app)
    popup.grab_set()

    label = ctk.CTkLabel(popup, text=message, font=("Arial", 14))
    label.pack(pady=20)

    def on_yes():
        result[0] = True
        popup.destroy()

    def on_no():
        result[0] = False
        popup.destroy()

    yes_button = ctk.CTkButton(popup, text="Oui", command=on_yes)
    yes_button.pack(side="left", padx=20, pady=10)

    no_button = ctk.CTkButton(popup, text="Non", command=on_no)
    no_button.pack(side="right", padx=20, pady=10)

    popup.wait_window()
    return result[0]


# ----------------------------------------------------------------------------------------

btn_Frame = ctk.CTkFrame(frame_1, width=300, height=200)
btn_Frame.place(x=10, y=450)

button_1_image = resize_image("images/add-to-cart.png", (30, 30))
button1 = ctk.CTkButton(
    btn_Frame,
    text="Ajouter",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=button_1_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    command=add_produit,
)
button1.place(x=10, y=10)

button_2_image = resize_image("images/edit.png", (30, 30))
button2 = ctk.CTkButton(
    btn_Frame,
    text="Modifiert",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=button_2_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    fg_color="#FF8C00",
    hover_color="#FF7000",
    command=update_produit,
)
button2.place(x=160, y=10)

button_3_image = resize_image("images/clean.png", (30, 30))
button3 = ctk.CTkButton(
    btn_Frame,
    text="Clean",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=button_3_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    fg_color="#2196F3",
    hover_color="#1565C0",
    command=clear_fields,
)
button3.place(x=10, y=55)

button_4_image = resize_image("images/trash.png", (30, 30))
button4 = ctk.CTkButton(
    btn_Frame,
    text="Supprim",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=button_4_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    fg_color="#e74c3c",
    hover_color="#c0392b",
    command=delete_produit,
)
button4.place(x=160, y=55)
# ===================deux frame ==========================================
tableview = ctk.CTkTabview(app, width=930, height=500)
tableview.place(x=370, y=190)
tableview.add("tableau")
tableview.add("graphique")


def draw_chart(chart_type):
    for widget in frame_graph.winfo_children():
        widget.destroy()

    data = db.fetch()
    if not data:
        return

    categories = [row[5] for row in data]
    quantities = [row[2] for row in data]

    fig = Figure(figsize=(5, 3), dpi=100)
    ax = fig.add_subplot(111)

    if chart_type == "Bar":
        ax.bar(categories, quantities, color="skyblue")
        ax.set_title("Quantité par Catégorie")
        ax.set_ylabel("Quantité")
        ax.set_xlabel("Catégorie")
    elif chart_type == "Pie":
        category_counts = {}
        for cat, qty in zip(categories, quantities):
            category_counts[cat] = category_counts.get(cat, 0) + qty
        ax.pie(
            category_counts.values(), labels=category_counts.keys(), autopct="%1.1f%%"
        )
        ax.set_title("Répartition des Quantités")

    canvas = FigureCanvasTkAgg(fig, master=frame_graph)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=ctk.BOTH, expand=True)


frame_graph = ttk.Frame(tableview.tab("graphique"), borderwidth=2, relief="groove")
frame_graph.pack(fill=ctk.BOTH, expand=True, padx=5, pady=5)


frame_controls = ttk.Frame(tableview.tab("graphique"))
frame_controls.pack(fill=ctk.X)


btn_bar_image = resize_image("images/bar-chart.png", (30, 30))
btn_bar = ctk.CTkButton(
    frame_controls,
    text="Afficher Bar Chart",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=btn_bar_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    command=lambda: draw_chart("Bar"),
)
btn_bar.pack(side=ctk.LEFT, padx=5)

btn_pie_image = resize_image("images/pie-chart.png", (30, 30))
btn_pie = ctk.CTkButton(
    frame_controls,
    text="Afficher Pie Chart",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=btn_pie_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    command=lambda: draw_chart("Pie"),
)
btn_pie.pack(side=ctk.LEFT, padx=5)


# =========================================================================
def export_to_excel():
    data = db.fetch()
    if not data:
        show_warning("Aucune donnée à exporter!")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
    )
    if not file_path:
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    headers = ["ID", "Nom", "Quantité", "Prix", "Date", "Catégorie", "Commentaire"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(file_path)
    show_success(f"Le fichier a été enregistré avec succès à {file_path}")


btn_export_image = resize_image("images/excel.png", (30, 30))
btn_export = ctk.CTkButton(
    btn_Frame,
    text="To Excel",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=btn_export_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    command=export_to_excel,
)
btn_export.place(x=10, y=100)


# =================fake data to test the app==================================================
def display_data():
    for row in tv.get_children():
        tv.delete(row)
    for row in db.fetch():
        tv.insert("", "end", values=row)


def generate_fake_data(db, count=100):
    fake = Faker()
    categories = ["Électronique", "Vêtements", "Meubles", "Alimentation", "Jardinage"]
    for _ in range(count):
        nom = fake.word()
        quantite = fake.random_int(min=1, max=100)
        prix = f"{fake.random_number(digits=2)}.{fake.random_number(digits=2)}"
        date = fake.date()
        categorie = fake.random.choice(categories)
        commentaire = fake.sentence()
        db.insert(nom, quantite, prix, date, categorie, commentaire)


def generate_and_display_data():
    if not db.is_empty():
        if show_confirm("Les données existent déjà. Voulez-vous le remplacer ?"):
            db.clear()
        else:
            return
    generate_fake_data(db)
    display_data()


fake_btn_image = resize_image("images\\fake.png", (30, 30))
fake_btn = ctk.CTkButton(
    btn_Frame,
    text="Fake data",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=fake_btn_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    command=generate_and_display_data,
)
fake_btn.place(x=160, y=100)


def delete_all():
    if show_confirm(
        "Êtes-vous sûr de vouloir effacer toutes les données de la base de données ?"
    ):
        db.clear()
        displayAll()


clear_all_image = resize_image("images\eraser.png", (30, 30))
clear_all = ctk.CTkButton(
    btn_Frame,
    text="Clear all",
    width=130,
    height=35,
    font=("Arial", 15, "bold"),
    cursor="hand2",
    image=clear_all_image,
    compound="right",
    corner_radius=10,
    border_width=1,
    fg_color="#e74c3c",
    hover_color="#c0392b",
    command=delete_all,
)
clear_all.place(x=10, y=145)


# ---------------------------------------------------------------------------------------------------
# =========================================searche=============================================


def update_table(data=None):
    for row in tv.get_children():
        tv.delete(row)
    if data is None:
        data = db.fetch()
    for row in data:
        tv.insert("", "end", values=row)


def search():
    column = search_by.get()
    search_value = search_entry.get()
    if not column or not search_value:
        show_warning(
            "Veuillez sélectionner une colonne et entrer une valeur pour la recherche !"
        )
        return

    query = f"SELECT * FROM Produits WHERE {column} LIKE ?"
    data = db.cur.execute(query, (f"%{search_value}%",)).fetchall()
    update_table(data)


def reset_table():
    search_entry.delete(0, ctk.END)
    update_table()


frame_search = ctk.CTkFrame(app)
# frame_search.pack(fill=ctk.Y, padx=5, pady=5)
frame_search.place(x=370, y=160)

search_by_label = ctk.CTkLabel(frame_search, text="Rechercher par:")
search_by_label.pack(side=ctk.LEFT, padx=5)

search_by = ctk.CTkComboBox(
    frame_search, values=["Nom", "Date", "Categorie", "Commentaire"], state="readonly"
)
search_by.pack(side=ctk.LEFT, padx=5)

search_entry = ctk.CTkEntry(frame_search)
search_entry.pack(side=ctk.LEFT, padx=5)

search_button = ctk.CTkButton(frame_search, text="Rechercher", command=search)
search_button.pack(side=ctk.LEFT, padx=5)

reset_button = ctk.CTkButton(frame_search, text="Réinitialiser", command=reset_table)
reset_button.pack(side=ctk.LEFT, padx=5)


# =============================================================================================

tableau_frame = ctk.CTkFrame(
    tableview.tab("tableau"),
    width=930,
    height=400,
    corner_radius=20,
)
tableau_frame.place(x=0, y=0)

style = ttk.Style()
style.configure("mystyle.Treeview", font=("Calibri", 13), rowheigh=50)
style.configure("mystyle.Treeview.Heading", font=("Calibri", 13))

tv = ttk.Treeview(
    tableau_frame, columns=(1, 2, 3, 4, 5, 6, 7), style="mystyle.Treeview"
)
tv.heading("1", text="ID")
tv.column("1", width=200)

tv.heading("2", text="Nom")
tv.column("2", width=100)

tv.heading("3", text="Quantite")
tv.column("3", width=120)

tv.heading("4", text="Prix")
tv.column("4", width=170)

tv.heading("5", text="Date")
tv.column("5", width=200)

tv.heading("6", text="Categorie")
tv.column("6", width=150)

tv.heading("7", text="Commentaire")
tv.column("7", width=230)
tv["show"] = "headings"
tv.bind("<ButtonRelease-1>", getData)
tv.place(x=10, y=10, width=1140, height=600)


displayAll()
app.mainloop()
